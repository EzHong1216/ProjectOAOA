import requests
from openpyxl import Workbook

requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)


class bcolors:
    HEADER = "\033[95m"
    OKBLUE = "\033[94m"
    OKCYAN = "\033[96m"
    OKGREEN = "\033[92m"
    WARNING = "\033[93m"
    FAIL = "\033[91m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"
    UNDERLINE = "\033[4m"


write_wb = Workbook()
write_ws = write_wb.active
write_ws.cell(1, 1, "번호")
write_ws.cell(1, 2, "IP")
write_ws.cell(1, 3, "결과")
write_ws.cell(1, 4, "비고")

f = open("./list.txt", 'r')

no = 0

print("================진단 중================")

while True:

    try:
        # IP 목록
        line = f.readline().rstrip()
        if not line: break
        no += 1

        # Request 요청
        response = requests.get(url=line, timeout=3, verify=False)

        # IF 구문작성을 위한 특정사이트 내용 확인
        # if no==22:
        #     print(line+response.url)
        # 분류 및 예외처리 if 문

        if "Please enable JavaScript" in response.text:
            print(str(no) + " " + bcolors.WARNING + line + ": 수동점검 - Enable JS" + bcolors.ENDC)
            # warningSite.append = no
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "수동점검")
            write_ws.cell(no + 1, 4, "Enable JS")

        if "리다이렉션 되는 URL 넣어주세요" in response.url:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(리다이렉션)" + bcolors.ENDC)
            # warningSite.append = no
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출(리다이렉션)")
        else:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호(해당 없음) - 페이지 노출(리다이렉션)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호(해당 없음)")
            write_ws.cell(no + 1, 4, "페이지 노출(리다이렉션)")

        if "보안정책에 따라 차단" in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - 보안정책 페이지" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "보안정책 페이지로 접근제어")

        if "invalid page fault" in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - 기타 웹 에러 페이지" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "기타 웹 에러페이지")

        if "error page test" in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - 별도의 접근제어 페이지(error page test)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "별도의 접근제어 페이지")

        if "Error 404" in response.text:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(WEB 서버 404 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출(WAS 404 페이지)")
        else:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호(해당 없음) - 페이지 노출(WEB 서버 404 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호(해당 없음)")
            write_ws.cell(no + 1, 4, "페이지 노출(WAS 404 페이지)")

        if "Wrong approach path" in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - 별도의 접근제어 페이지(Wrong approach path)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "별도의 접근제어 페이지(Wrong approach path)")

        if "/error/error_img" in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - 별도의 접근제어 페이지(/error/error_img)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "별도의 접근제어 페이지(/error/error_img)")

        if "Error 403</h2>" in response.text:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(IIS 403 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출(IIS 403 페이지)")
        else:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호(해당 없음) - 페이지 노출(IIS 403 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호(해당 없음)")
            write_ws.cell(no + 1, 4, "페이지 노출(IIS 403 페이지)")

        if "<h2>404 -" in response.text:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(IIS 404 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출(IIS 404 페이지)")
        else:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호(해당 없음) - 페이지 노출(IIS 404 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호(해당 없음)")
            write_ws.cell(no + 1, 4, "페이지 노출(IIS 404 페이지)")

        if "Invalid" in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - 별도의 접근제어 페이지" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "별도의 접근제어 페이지")

        if "You are unauthorized to access this page." in response.text:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호 - WAS 403 접근제어 페이지" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호")
            write_ws.cell(no + 1, 4, "WAS 403 접근제어 페이지")

        if "you've successfully installed Tomcat" in response.text:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(Apache 기본 페이지)" + bcolors.ENDC)
            # warningSite.append = no
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출(Apache 기본 페이지)")
        else:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호(해당 없음) - 페이지 노출(Apache 기본 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호(해당 없음)")
            write_ws.cell(no + 1, 4, "페이지 노출(Apache 기본 페이지)")

        if "IIS Windows Server" in response.text:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(IIS 기본 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출(IIS 기본 페이지)")
        else:
            print(str(no) + " " + bcolors.OKGREEN + line + ": 양호(해당 없음) - 페이지 노출(IIS 기본 페이지)" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "양호(해당 없음)")
            write_ws.cell(no + 1, 4, "페이지 노출(IIS 기본 페이지)")


        if "아이디" in response.text or "password" in response.text or "login" in response.text:
            if "https" in response.url:
                print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출" + bcolors.ENDC)
                write_ws.cell(no + 1, 1, str(no))
                write_ws.cell(no + 1, 2, line)
                write_ws.cell(no + 1, 3, "취약")
                write_ws.cell(no + 1, 4, "페이지 노출")

            else:
                print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출(로그인 http)" + bcolors.ENDC)
                write_ws.cell(no + 1, 1, str(no))
                write_ws.cell(no + 1, 2, line)
                write_ws.cell(no + 1, 3, "취약")
                write_ws.cell(no + 1, 4, "페이지 노출(로그인 http)")
        else:
            print(str(no) + " " + bcolors.FAIL + line + ": 취약 - 페이지 노출" + bcolors.ENDC)
            write_ws.cell(no + 1, 1, str(no))
            write_ws.cell(no + 1, 2, line)
            write_ws.cell(no + 1, 3, "취약")
            write_ws.cell(no + 1, 4, "페이지 노출")

    except requests.exceptions.Timeout:
        print(str(no) + " " + bcolors.OKCYAN + line + ": 양호 - 예외처리(응답없음)" + bcolors.ENDC)
        write_ws.cell(no + 1, 1, str(no))
        write_ws.cell(no + 1, 2, line)
        write_ws.cell(no + 1, 3, "양호")
        write_ws.cell(no + 1, 4, "예외처리(타임아웃)")
        pass

    # SSL 인증서 에러 발생 방지
    except requests.exceptions.SSLError:
        print(str(no) + " " + bcolors.OKCYAN + line + ": 양호 - 예외처리(SSLError)" + bcolors.ENDC)
        write_ws.cell(no + 1, 1, str(no))
        write_ws.cell(no + 1, 2, line)
        write_ws.cell(no + 1, 3, "양호")
        write_ws.cell(no + 1, 4, "예외처리(SSLError)")
        pass

    # 연결 실패(응답 없음) 예외처리
    except requests.exceptions.ConnectionError:
        print(str(no) + " " + bcolors.OKCYAN + line + ": 양호 - 예외처리(응답없음)" + bcolors.ENDC)
        write_ws.cell(no + 1, 1, str(no))
        write_ws.cell(no + 1, 2, line)
        write_ws.cell(no + 1, 3, "양호")
        write_ws.cell(no + 1, 4, "예외처리(응답없음)")
        pass





f.close()
write_wb.save('./result.xlsx')
