import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
import urllib.request, urllib.error, urllib.parse
import pprint

requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)


class bcolors:
    HEADER = "\033[95m"
    OKBLUE = "\033[94m"
    OKCYAN = "\033[96m"
    OKGREEN = "\033[92m"
    WARNING = "\033[93m"
    FAIL = "\033[91m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"
    UNDERLINE = "\033[4m"

class web_weak_scanner:
    def __init__(self):
        self.wb = Workbook()
        self.reset_WB()
    
    def reset_WB(self):
        self.wb.close()
        self.wb = Workbook()
        ws = self.wb.active
        ws.cell(1, 1, "번호")
        ws.cell(1, 2, "IP")
        ws.cell(1, 3, "결과")
        ws.cell(1, 4, "비고")
        self.no = 1
    
    def append_Data(self, url, result, etc):
        ws = self.wb.active
        ws.cell(self.no + 1, 1, str(self.no))
        ws.cell(self.no + 1, 2, url)
        ws.cell(self.no + 1, 3, result)
        ws.cell(self.no + 1, 4, etc)
        self.no += 1

    #사이트 노출 취약 스캔

    def scan(self, url_list: list):
        self.reset_WB()
        result = []
        for url in url_list:
            url = url.rstrip()
            try:
                # Request 요청
                response = requests.get(url=url, timeout=3, verify=False)

                # IF 구문작성을 위한 특정사이트 내용 확인
                # if no==22:
                #     print(line+response.url)
                # 분류 및 예외처리 if 문

                if "Please enable JavaScript" in response.text:
                    print(str(self.no) + " " + bcolors.WARNING + url + ": 수동점검필요 - Enable JS" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 수동점검필요 - Enable JS")
                    
                    # warningSite.append = no
                    self.append_Data(url, "수동점검", "Enable JS")

                if "리다이렉션 되는 URL 넣어주세요" in response.url:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(리다이렉션)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(리다이렉션)")
                    # warningSite.append = no
                    self.append_Data(url, "취약", "페이지 노출(리다이렉션)")
                else:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호(해당 없음) - 페이지 노출(리다이렉션)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호(해당 없음) - 페이지 노출(리다이렉션)")
                    self.append_Data(url, "양호(해당 없음)", "페이지 노출(리다이렉션)")

                if "보안정책에 따라 차단" in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - 보안정책 페이지" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - 보안정책 페이지")
                    self.append_Data(url, "양호", "보안정책 페이지로 접근제어")
                    
                if "invalid page fault" in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - 기타 웹 에러 페이지" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - 기타 웹 에러 페이지")
                    self.append_Data(url, "양호", "기타 웹 에러페이지")

                if "error page test" in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - 별도의 접근제어 페이지(error page test)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - 별도의 접근제어 페이지(error page test)")
                    self.append_Data(url, "양호", "별도의 접근제어 페이지")

                if "Error 404" in response.text:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(WEB 서버 404 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(WEB 서버 404 페이지)")
                    self.append_Data(url, "취약", "페이지 노출(WAS 404 페이지)")
                else:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호(해당 없음) - 페이지 노출(WEB 서버 404 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호(해당 없음) - 페이지 노출(WEB 서버 404 페이지)")
                    self.append_Data(url, "양호(해당 없음)", "페이지 노출(WAS 404 페이지)")

                if "Wrong approach path" in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - 별도의 접근제어 페이지(Wrong approach path)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - 별도의 접근제어 페이지(Wrong approach path)")
                    self.append_Data(url, "양호", "별도의 접근제어 페이지(Wrong approach path)")

                if "/error/error_img" in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - 별도의 접근제어 페이지(/error/error_img)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - 별도의 접근제어 페이지(/error/error_img)")
                    self.append_Data(url, "양호", "별도의 접근제어 페이지(/error/error_img)")

                if "Error 403</h2>" in response.text:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(IIS 403 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(IIS 403 페이지)")
                    self.append_Data(url, "취약", "페이지 노출(IIS 403 페이지)")
                else:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호(해당 없음) - 페이지 노출(IIS 403 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호(해당 없음) - 페이지 노출(IIS 403 페이지)")
                    self.append_Data(url, "양호(해당 없음)", "페이지 노출(IIS 403 페이지)")

                if "<h2>404 -" in response.text:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(IIS 404 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(IIS 404 페이지)")
                    self.append_Data(url, "취약", "페이지 노출(IIS 404 페이지)")
                else:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호(해당 없음) - 페이지 노출(IIS 404 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호(해당 없음) - 페이지 노출(IIS 404 페이지)")
                    self.append_Data(url, "양호(해당 없음)", "페이지 노출(IIS 404 페이지)")
                    
                if "Invalid" in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - 별도의 접근제어 페이지" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - 별도의 접근제어 페이지")
                    self.append_Data(url, "양호", "별도의 접근제어 페이지")

                if "You are unauthorized to access this page." in response.text:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호 - WAS 403 접근제어 페이지" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호 - WAS 403 접근제어 페이지")
                    self.append_Data(url, "양호", "WAS 403 접근제어 페이지")

                if "you've successfully installed Tomcat" in response.text:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(Apache 기본 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(Apache 기본 페이지)")
                    # warningSite.append = no
                    self.append_Data(url, "취약", "페이지 노출(Apache 기본 페이지)")
                else:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호(해당 없음) - 페이지 노출(Apache 기본 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호(해당 없음) - 페이지 노출(Apache 기본 페이지)")
                    self.append_Data(url, "양호(해당 없음)", "페이지 노출(Apache 기본 페이지)")

                if "IIS Windows Server" in response.text:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(IIS 기본 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(IIS 기본 페이지)")
                    self.append_Data(url, "취약", "페이지 노출(IIS 기본 페이지)")
                else:
                    print(str(self.no) + " " + bcolors.OKGREEN + url + ": 양호(해당 없음) - 페이지 노출(IIS 기본 페이지)" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 양호(해당 없음) - 페이지 노출(IIS 기본 페이지)")
                    self.append_Data(url, "양호(해당 없음)", "페이지 노출(IIS 기본 페이지)")


                if "아이디" in response.text or "password" in response.text or "login" in response.text:
                    if "https" in response.url:
                        print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출" + bcolors.ENDC)
                        result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출")
                        self.append_Data(url, "취약", "페이지 노출")
                    else:
                        print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출(로그인 http)" + bcolors.ENDC)
                        result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출(로그인 http)")
                        self.append_Data(url, "취약", "페이지 노출(로그인 http)")
                else:
                    print(str(self.no) + " " + bcolors.FAIL + url + ": 취약 - 페이지 노출" + bcolors.ENDC)
                    result.append(str(self.no) + " " + url + ": 취약 - 페이지 노출")
                    self.append_Data(url, "취약", "페이지 노출")

            except requests.exceptions.Timeout:
                print(str(self.no) + " " + bcolors.OKCYAN + url + ": 양호 - 예외처리(응답없음)" + bcolors.ENDC)
                result.append(str(self.no) + " " + url + ": 양호 - 예외처리(응답없음)")
                self.append_Data(url, "양호", "예외처리(타임아웃)")
                pass

            # SSL 인증서 에러 발생 방지
            except requests.exceptions.SSLError:
                print(str(self.no) + " " + bcolors.OKCYAN + url + ": 양호 - 예외처리(SSLError)" + bcolors.ENDC)
                result.append(str(self.no) + " " + url + ": 양호 - 예외처리(SSLError)")
                self.append_Data(url, "양호", "예외처리(SSLError)")
                pass

            # 연결 실패(응답 없음) 예외처리
            except requests.exceptions.ConnectionError:
                print(str(self.no) + " " + bcolors.OKCYAN + url + ": 양호 - 예외처리(응답없음)" + bcolors.ENDC)
                result.append(str(self.no) + " " + url + ": 양호 - 예외처리(SSLError)")
                self.append_Data(url, "양호", "예외처리(응답없음)")
                pass
            
        return result

    def save_Excel(self, directory: str):
        self.wb.save(directory)

# 주소 한번만 나오게 수정
# 리턴값 추가

# 취약점 스캔용 리스트 수정 가능한 txt 파일 열기 기능 gui 필요 - 새 창
# 메인화면에서 버튼 실행시 새 창 띄워서 진행 할 수 있도록 .
# 새 창